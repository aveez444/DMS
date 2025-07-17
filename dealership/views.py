from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.models import Group
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.conf import settings
import logging
from django.utils import timezone
from datetime import timedelta
from django.http import Http404
from accounts.models import CustomUser
from tenants.models import Client, Domain
from django_tenants.utils import tenant_context
from threading import local
import jwt
import openpyxl
from rest_framework import generics
from django.db.models import Sum
from datetime import datetime, timedelta
from .serializers import ( CombinedVehicleSerializer,CatalogueSerializer, VehicleImageSerializer, OutboundVehicleSerializer, UpdateCatalogueSerializer, PaymentSerializer, VehicleListSerializer,VehicleDetailSerializer,VehicleInquirySerializer, InquiryBrokerSerializer, VehicleDataSerializer,OutboundVehicleSerializer, MaintenanceRecord, MaintenanceRecordSerializer,VehicleUpdateSerializer,StaffSerializer,StaffSalarySerializer,StaffSalaryMonthWiseSerializer,InvoiceSerializer,ElectricityBillSerializer, OfficeRentSerializer, WifiBillSerializer, AdditionalExpenseSerializer)

from .models import Vehicle, MaintenanceRecord, VehicleImage, OutboundVehicle, Payment, Inquirybroker, VehicleInquiry, Staff,StaffSalary,Invoice,ElectricityBill, OfficeRent, WifiBill, AdditionalExpense

logger = logging.getLogger(__name__)


class CombinedVehicleAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        if not (request.user.has_perm('dealership.add_vehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to add vehicles.'}, 
                          status=status.HTTP_403_FORBIDDEN)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating vehicle for tenant: {tenant.name}")  # Fixed typo: tenant6565 -> tenant.name
                serializer = CombinedVehicleSerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    vehicle = serializer.save()
                    return Response({"message": "Vehicle added", "vehicle_id": vehicle.vehicle_id}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in CombinedVehicleAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)


            
class VehicleImageAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def get(self, request, vehicle_id):
        if not request.user.is_authenticated:
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        try:
            with tenant_context(request.tenant):
                vehicle = Vehicle.objects.get(vehicle_id=vehicle_id)
                images = vehicle.images.all()
                serializer = VehicleImageSerializer(images, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Vehicle.DoesNotExist:
            return Response({"error": "Vehicle not found"}, status=404)
        except Exception as e:
            logger.error(f"Error in VehicleImageAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def post(self, request, vehicle_id):
        if not request.user.is_authenticated:
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        if not (request.user.has_perm('dealership.add_vehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to add vehicle images.'}, 
                          status=status.HTTP_403_FORBIDDEN)

        try:
            with tenant_context(request.tenant):
                vehicle = Vehicle.objects.get(vehicle_id=vehicle_id)
                images_data = request.FILES.getlist('images')
                if not images_data:
                    return Response({"error": "No images provided"}, status=400)
                
                image_instances = []
                for image_data in images_data:
                    image = VehicleImage(vehicle=vehicle, image=image_data)
                    image_instances.append(image)
                
                VehicleImage.objects.bulk_create(image_instances)
                serializer = VehicleImageSerializer(image_instances, many=True, context={"request": request})
                return Response({"message": "Images uploaded", "images": serializer.data}, status=201)
        except Vehicle.DoesNotExist:
            return Response({"error": "Vehicle not found"}, status=404)
        except Exception as e:
            logger.error(f"Error in VehicleImageAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)


class VehicleUpdateAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)
    
    def get_object(self, vehicle_id, tenant):
        with tenant_context(tenant):
            return get_object_or_404(Vehicle, vehicle_id=vehicle_id)
    
    def put(self, request, vehicle_id):
        """
        Full update of vehicle - requires all fields
        """
        logger.debug(f"PUT request to update vehicle {vehicle_id}")
        
        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)
            
        # Check permissions
        if not (request.user.has_perm('dealership.change_vehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to update vehicles.'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        tenant = request.tenant
        try:
            vehicle = self.get_object(vehicle_id, tenant)
            
            with tenant_context(tenant):
                serializer = VehicleUpdateSerializer(vehicle, data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Vehicle updated successfully"}, status=status.HTTP_200_OK)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error updating vehicle: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def patch(self, request, vehicle_id):
        """
        Partial update of vehicle - only specified fields
        """
        logger.debug(f"PATCH request to update vehicle {vehicle_id}")
        logger.debug(f"Request data: {request.data}")
        
        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)
            
        # Check permissions
        if not (request.user.has_perm('dealership.change_vehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to update vehicles.'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        tenant = request.tenant
        try:
            vehicle = self.get_object(vehicle_id, tenant)
            
            with tenant_context(tenant):
                # Special handling for files if they're present
                data = request.data.copy()
                
                # Handle file uploads if present
                file_fields = ['proof_of_ownership_document', 'upload_image_of_vehicle', 
                               'proof_of_ownership', 'purchase_agreement']
                
                for field in file_fields:
                    # If the field is not in the data or is an empty string, remove it
                    if field in data and (data[field] == '' or data[field] == 'null'):
                        data.pop(field)
                
                serializer = VehicleUpdateSerializer(vehicle, data=data, partial=True, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Vehicle updated successfully"}, status=status.HTTP_200_OK)
                
                logger.error(f"Validation errors: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error updating vehicle: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

logger = logging.getLogger(__name__)
class VehicleListView(APIView):

    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching vehicles for tenant: {tenant.name}")
                vehicles = Vehicle.objects.all()
                serializer = CombinedVehicleSerializer(vehicles, many=True, context={"request": request})
                return Response(serializer.data)
        except Exception as e:
            logger.error(f"Error in VehicleListView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
            
 
logger = logging.getLogger(__name__)

class LiveInventoryView(APIView):
      
    def get(self, request):
            logger.debug(f"Request headers: {request.headers}")
            logger.debug(f"Session: {dict(request.session)}")
            logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
            logger.debug(f"Cookies: {request.COOKIES}")

            if not request.user.is_authenticated:
                logger.warning("Unauthorized access attempt: No active session")
                return Response({"error": "Authentication credentials were not provided."}, status=401)

            tenant = request.tenant
            logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

            try:
                with tenant_context(tenant):
                    logger.debug(f"Fetching vehicles for tenant: {tenant.name}")
                    vehicles = Vehicle.objects.all()
                    serializer = CombinedVehicleSerializer(vehicles, many=True, context={"request": request})
                    return Response(serializer.data)
            except Exception as e:
                logger.error(f"Error in LiveInventoryView: {str(e)}", exc_info=True)
                return Response({"error": f"Server error: {str(e)}"}, status=500)
                


class DeleteVehicleAPIView(APIView):
    def delete(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        # Check if user has permission to delete vehicles or belongs to allowed groups
        if not (request.user.has_perm('dealership.delete_vehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to delete vehicles.'}, 
                          status=status.HTTP_403_FORBIDDEN)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Deleting vehicle for tenant: {tenant.name}")
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                vehicle.delete()
                return Response({"message": f"Vehicle with ID {vehicle_id} deleted successfully."}, status=200)
        except Exception as e:
            logger.error(f"Error in DeleteVehicleAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)


class VehicleDataAPIView(APIView):
    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching vehicles for tenant: {tenant.name}")
                vehicles = Vehicle.objects.all()
                serializer = VehicleDataSerializer(vehicles, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in VehicleDataAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)


class AddMaintenanceAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Adding maintenance record for tenant: {tenant.name}")
                vehicle_id = request.data.get('vehicle_id')
                if not vehicle_id:
                    return Response({"error": "Vehicle ID is required."}, status=400)

                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                maintenance_data = {
                    "vehicle_id": vehicle_id,
                    "maintenance_type": request.data.get("maintenance_type", "General Service"),
                    "maintenance_date": request.data.get("maintenance_date", datetime.now().date()),
                    "cost": request.data.get("cost", 0.0),
                    "person_in_charge": request.data.get("person_in_charge", "Not Provided"),
                    "receipt": request.FILES.get("receipt")
                }

                serializer = MaintenanceRecordSerializer(data=maintenance_data)
                if serializer.is_valid():
                    serializer.save()
                    return Response(
                        {"message": "Maintenance added successfully!", "data": serializer.data},
                        status=201
                    )
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in AddMaintenanceAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
            
class MaintenanceRecordListCreateView(APIView):
    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching maintenance records for tenant: {tenant.name}")
                records = MaintenanceRecord.objects.all().order_by('-created_at')
                serializer = MaintenanceRecordSerializer(records, many=True, context={'request': request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in MaintenanceRecordListCreateView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating maintenance record for tenant: {tenant.name}")
                serializer = MaintenanceRecordSerializer(data=request.data, context={'request': request})
                if serializer.is_valid():
                    serializer.save()
                    return Response(serializer.data, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in MaintenanceRecordListCreateView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class MaintenanceRecordDetailView(APIView):
    def get(self, request, pk):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching maintenance record for tenant: {tenant.name}")
                record = get_object_or_404(MaintenanceRecord, pk=pk)
                serializer = MaintenanceRecordSerializer(record, context={'request': request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in MaintenanceRecordDetailView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def put(self, request, pk):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Updating maintenance record for tenant: {tenant.name}")
                record = get_object_or_404(MaintenanceRecord, pk=pk)
                serializer = MaintenanceRecordSerializer(record, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response(serializer.data, status=200)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in MaintenanceRecordDetailView PUT: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def delete(self, request, pk):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Deleting maintenance record for tenant: {tenant.name}")
                record = get_object_or_404(MaintenanceRecord, pk=pk)
                record.delete()
                return Response({"message": "Record deleted successfully."}, status=204)
        except Exception as e:
            logger.error(f"Error in MaintenanceRecordDetailView DELETE: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

logger = logging.getLogger(__name__)

class MaintenanceRecordExportView(APIView):
    """
    Export all maintenance records as an Excel file.
    """
    def get(self, request):
        # 1) Create an in-memory workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maintenance Records"

        # 2) Write column headers (adjust as needed)
        headers = [
            "ID",
            "Vehicle ID",
            "Maintenance Type",
            "Maintenance Date",
            "Cost",
            "Person in Charge",
            "Receipt",
            "Created At",
        ]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # 3) Query all maintenance records
        queryset = MaintenanceRecord.objects.all().order_by('id')

        # 4) Write each record row
        row_index = 2
        for record in queryset:
            ws.cell(row=row_index, column=1, value=record.id)
            ws.cell(row=row_index, column=2, value=record.vehicle.vehicle_id)  # or record.vehicle_id
            ws.cell(row=row_index, column=3, value=record.maintenance_type)
            # Convert date to string if not None
            ws.cell(row=row_index, column=4, value=str(record.maintenance_date) if record.maintenance_date else "")
            ws.cell(row=row_index, column=5, value=record.cost)
            ws.cell(row=row_index, column=6, value=record.person_in_charge if record.person_in_charge else "")
            ws.cell(row=row_index, column=7, value=str(record.receipt) if record.receipt else "No Receipt")
            ws.cell(row=row_index, column=8, value=str(record.created_at) if record.created_at else "")
            row_index += 1

        # 5) Prepare HTTP response with an Excel attachment
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=maintenance_records.xlsx'

        # 6) Save the workbook content to the response
        wb.save(response)
        return response


class VehicleInventoryExportView(APIView):
    def get(self, request):
        # create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vehicle Inventory"

        # headers
        headers = [
            "Vehicle ID", "Type", "Make", "Model", "Year",
            "Chassis", "Plate", "Odometer", "Color", "Fuel",
            "Transmission", "Seller", "Mobile", "Email",
            "Condition", "Tires", "Damage", "Engine",
            "Interior", "Arrival", "Price", "Storage", "Notes"
        ]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # query vehicles
        queryset = Vehicle.objects.all().order_by("vehicle_id")

        # fill rows
        row_index = 2
        for v in queryset:
            ws.cell(row=row_index, column=1, value=v.vehicle_id)
            ws.cell(row=row_index, column=2, value=v.vehicle_type)
            ws.cell(row=row_index, column=3, value=v.vehicle_make)
            ws.cell(row=row_index, column=4, value=v.vehicle_model)
            ws.cell(row=row_index, column=5, value=v.year_of_manufacturing)
            ws.cell(row=row_index, column=6, value=v.chassis_number)
            ws.cell(row=row_index, column=7, value=v.license_plate_number)
            ws.cell(row=row_index, column=8, value=v.odometer_reading_kms)
            ws.cell(row=row_index, column=9, value=v.color)
            ws.cell(row=row_index, column=10, value=v.fuel_type)
            ws.cell(row=row_index, column=11, value=v.transmission_type)
            ws.cell(row=row_index, column=12, value=v.seller_name_company_name)
            ws.cell(row=row_index, column=13, value=v.mobile_number)
            ws.cell(row=row_index, column=14, value=v.email_address)
            ws.cell(row=row_index, column=15, value=v.condition_grade)
            ws.cell(row=row_index, column=16, value=v.tires_condition)
            ws.cell(row=row_index, column=17, value=v.damage_details_if_any)
            ws.cell(row=row_index, column=18, value=v.engine_condition)
            # if interior_condition is boolean or string
            ws.cell(row=row_index, column=19, value=str(v.interior_condition))
            ws.cell(row=row_index, column=20, value=str(v.arrival_date or ""))
            ws.cell(row=row_index, column=21, value=str(v.purchase_price or ""))
            ws.cell(row=row_index, column=22, value=v.storage_location)
            ws.cell(row=row_index, column=23, value=v.notes or "")
            row_index += 1

        # response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=vehicle_inventory.xlsx'
        wb.save(response)
        return response
        
class OutboundVehicleAPIView(APIView):
    def get(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        # Check if user has permission to view outbound vehicles or belongs to allowed groups
        if not (request.user.has_perm('dealership.view_outboundvehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to view outbound vehicles.'}, 
                          status=status.HTTP_403_FORBIDDEN)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching outbound vehicle record for tenant: {tenant.name}")
                try:
                    outbound_vehicle = get_object_or_404(OutboundVehicle, vehicle_id=vehicle_id)
                    serializer = OutboundVehicleSerializer(outbound_vehicle)
                    return Response(serializer.data, status=200)
                except Http404:
                    return Response(
                        {"message": f"No outbound record found for Vehicle ID {vehicle_id}"},
                        status=404
                    )
        except Exception as e:
            logger.error(f"Error in OutboundVehicleAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
    
    def post(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        # Check if user has permission to add outbound vehicles or belongs to allowed groups
        if not (request.user.has_perm('dealership.add_outboundvehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to add outbound vehicles.'}, 
                          status=status.HTTP_403_FORBIDDEN)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating outbound vehicle record for tenant: {tenant.name}")
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                data = request.data.copy()
                data["vehicle"] = vehicle_id
                serializer = OutboundVehicleSerializer(data=data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response(
                        {"message": "Outbound vehicle record created successfully!", "data": serializer.data},
                        status=201
                    )
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in OutboundVehicleAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)


class UpdateOutboundVehicleAPIView(APIView):
    def patch(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        # Check if user has permission to change outbound vehicles or belongs to allowed groups
        if not (request.user.has_perm('dealership.change_outboundvehicle') or 
                request.user.groups.filter(name__in=['sub-admins', 'salesperson']).exists()):
            return Response({'error': 'You do not have permission to update outbound vehicles.'}, 
                          status=status.HTTP_403_FORBIDDEN)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Updating outbound vehicle for tenant: {tenant.name}")
                outbound_vehicle = get_object_or_404(OutboundVehicle, vehicle_id=vehicle_id)
                serializer = OutboundVehicleSerializer(outbound_vehicle, data=request.data, partial=True, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response(
                        {"message": f"Outbound details updated successfully for Vehicle ID {vehicle_id}", "data": serializer.data},
                        status=200
                    )
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in UpdateOutboundVehicleAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class CreatePaymentAPIView(APIView):

    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating payments for tenant: {tenant.name}")
                vehicle_id = request.data.get("vehicle_id")
                if not vehicle_id:
                    return Response({"error": "Vehicle ID is required."}, status=400)

                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                payment_slots = request.data.get("payment_slots", [])
                if not payment_slots:
                    return Response({"error": "Payment slots array is required."}, status=400)

                created_payments = []
                for slot in payment_slots:
                    serializer = PaymentSerializer(data={**slot, "vehicle_id": vehicle_id}, context={"request": request})
                    if serializer.is_valid():
                        serializer.save()
                        created_payments.append(serializer.data)
                    else:
                        return Response(serializer.errors, status=400)

                return Response(
                    {"message": "Payments added successfully!", "data": created_payments},
                    status=201
                )
        except Exception as e:
            logger.error(f"Error in CreatePaymentAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class ViewPaymentsAPIView(APIView):
    def get(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching payments for tenant: {tenant.name}")
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                payments = Payment.objects.filter(vehicle=vehicle).order_by("slot_number")

                if not payments.exists():
                    return Response({"message": f"No payments found for vehicle ID {vehicle_id}."}, status=404)

                serializer = PaymentSerializer(payments, many=True, context={"request": request})
                return Response(serializer.data, status=200)

        except Exception as e:
            logger.error(f"Error in ViewPaymentsAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
  
class UpdatePaymentAPIView(APIView):
    def put(self, request, vehicle_id, slot_number):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Updating payment for tenant: {tenant.name}")
                payment_type = request.data.get("payment_type")
                if payment_type not in ["purchase", "selling"]:
                    return Response({"error": "Invalid payment type. Must be 'purchase' or 'selling'."}, status=400)

                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                payment_instance, created = Payment.objects.get_or_create(
                    vehicle=vehicle, slot_number=slot_number, payment_type=payment_type
                )

                serializer = PaymentSerializer(payment_instance, data=request.data, partial=True, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    message = "Payment updated successfully!" if not created else "New payment slot added!"
                    return Response({"message": message, "updated_data": serializer.data}, status=200)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in UpdatePaymentAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def patch(self, request, vehicle_id, slot_number):
        if not request.user.is_authenticated:
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        try:
            with tenant_context(tenant):
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                payment_type = request.data.get("payment_type")

                if not payment_type:
                    return Response({"error": "payment_type is required."}, status=400)

                payment_instance = get_object_or_404(
                    Payment,
                    vehicle=vehicle,
                    slot_number=slot_number,
                    payment_type=payment_type
                )

                serializer = PaymentSerializer(payment_instance, data=request.data, partial=True, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Payment partially updated.", "updated_data": serializer.data}, status=200)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in UpdatePaymentAPIView PATCH: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
  
class VehiclePaymentSummaryAPIView(APIView):
    def get(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching payment summary for tenant: {tenant.name}")
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                outbound_vehicle = OutboundVehicle.objects.filter(vehicle=vehicle).first()
                selling_price = outbound_vehicle.selling_price if outbound_vehicle else 0
                purchase_price = vehicle.purchase_price or 0

                total_purchase_paid = Payment.objects.filter(vehicle=vehicle, payment_type="purchase").aggregate(
                    total=Sum('amount_paid')
                )['total'] or 0
                purchase_balance = purchase_price - total_purchase_paid

                total_selling_received = Payment.objects.filter(vehicle=vehicle, payment_type="selling").aggregate(
                    total=Sum('amount_paid')
                )['total'] or 0
                selling_balance = selling_price - total_selling_received

                return Response({
                    "vehicle_id": vehicle_id,
                    "purchase_price": purchase_price,
                    "total_purchase_paid": total_purchase_paid,
                    "purchase_balance": purchase_balance,
                    "selling_price": selling_price,
                    "total_selling_received": total_selling_received,
                    "selling_balance": selling_balance
                }, status=200)
        except Exception as e:
            logger.error(f"Error in VehiclePaymentSummaryAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
   
class VehicleCostAPIView(APIView):
    def get(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching vehicle cost for tenant: {tenant.name}")
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                purchase_price = vehicle.purchase_price if vehicle.purchase_price else 0
                total_maintenance_cost = MaintenanceRecord.objects.filter(vehicle=vehicle).aggregate(
                    total_cost=Sum('cost')
                )['total_cost'] or 0
                total_cost = purchase_price + total_maintenance_cost

                return Response({
                    "vehicle_id": vehicle_id,
                    "purchase_price": purchase_price,
                    "total_maintenance_cost": total_maintenance_cost,
                    "total_cost": total_cost
                }, status=200)
        except Exception as e:
            logger.error(f"Error in VehicleCostAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
   
class CatalogueAPIView(APIView):
    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching catalogue for tenant: {tenant.name}")
                vehicles = Vehicle.objects.filter(inventory_status="IN")
                for vehicle in vehicles:
                    vehicle.refresh_from_db()
                serializer = CatalogueSerializer(vehicles, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in CatalogueAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)


class UpdateCatalogueAPIView(APIView):
    parser_classes = (JSONParser,)

    def patch(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Updating catalogue for tenant: {tenant.name}")
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                serializer = UpdateCatalogueSerializer(vehicle, data=request.data, partial=True, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        "message": "Catalogue updated successfully!",
                        "updated_data": serializer.data
                    }, status=200)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in UpdateCatalogueAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
            
class CatalogueDetailAPIView(APIView):
    def get(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching catalogue details for tenant: {tenant.name}")
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                serializer = CatalogueSerializer(vehicle, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in CatalogueDetailAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class VehicleDetailAPIView(APIView):
    def get(self, request, vehicle_id):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt")
            return Response({"detail": "Authentication credentials were not provided."}, status=status.HTTP_401_UNAUTHORIZED)

        tenant = getattr(request, 'tenant', None)
        if tenant is None:
            return Response({"detail": "Tenant not found."}, status=status.HTTP_400_BAD_REQUEST)

        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                vehicle = get_object_or_404(Vehicle, vehicle_id=vehicle_id)
                serializer = VehicleDetailSerializer(vehicle, context={"request": request})
                return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error in VehicleDetailAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class InquiryBrokerListCreateAPIView(generics.ListCreateAPIView):
    queryset = Inquirybroker.objects.all()
    serializer_class = InquiryBrokerSerializer

    def get(self, request, *args, **kwargs):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching inquiry brokers for tenant: {tenant.name}")
                return super().get(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error in InquiryBrokerListCreateAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def post(self, request, *args, **kwargs):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating inquiry broker for tenant: {tenant.name}")
                return super().post(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error in InquiryBrokerListCreateAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class VehicleInquiryListCreateAPIView(generics.ListCreateAPIView):
    queryset = VehicleInquiry.objects.all()
    serializer_class = VehicleInquirySerializer

    def get(self, request, *args, **kwargs):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching vehicle inquiries for tenant: {tenant.name}")
                return super().get(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error in VehicleInquiryListCreateAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def post(self, request, *args, **kwargs):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating vehicle inquiry for tenant: {tenant.name}")
                return super().post(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error in VehicleInquiryListCreateAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
   
class ElectricityBillAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating electricity bill for tenant: {tenant.name}")
                serializer = ElectricityBillSerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Electricity Bill Added!", "data": serializer.data}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in ElectricityBillAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching electricity bills for tenant: {tenant.name}")
                records = ElectricityBill.objects.all()
                serializer = ElectricityBillSerializer(records, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in ElectricityBillAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
 
class OfficeRentAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating office rent record for tenant: {tenant.name}")
                serializer = OfficeRentSerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Office Rent Added!", "data": serializer.data}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in OfficeRentAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching office rent records for tenant: {tenant.name}")
                records = OfficeRent.objects.all()
                serializer = OfficeRentSerializer(records, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in OfficeRentAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class WifiBillAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating WiFi bill for tenant: {tenant.name}")
                serializer = WifiBillSerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "WiFi Bill Added!", "data": serializer.data}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in WifiBillAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching WiFi bills for tenant: {tenant.name}")
                records = WifiBill.objects.all()
                serializer = WifiBillSerializer(records, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in WifiBillAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
 
class AdditionalExpenseAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating additional expense for tenant: {tenant.name}")
                serializer = AdditionalExpenseSerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Additional Expense Added!", "data": serializer.data}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in AdditionalExpenseAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching additional expenses for tenant: {tenant.name}")
                records = AdditionalExpense.objects.all()
                serializer = AdditionalExpenseSerializer(records, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in AdditionalExpenseAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
  
class StaffAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating staff record for tenant: {tenant.name}")
                serializer = StaffSerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Staff added successfully!", "data": serializer.data}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in StaffAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching staff records for tenant: {tenant.name}")
                staff = Staff.objects.all()
                serializer = StaffSerializer(staff, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in StaffAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
   
class StaffSalaryAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating staff salary record for tenant: {tenant.name}")
                serializer = StaffSalarySerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Salary record added successfully!", "data": serializer.data}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in StaffSalaryAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)
 
class StaffSalaryUpdateAPIView(APIView):
    def put(self, request, pk):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Updating staff salary record for tenant: {tenant.name}")
                record = get_object_or_404(StaffSalary, pk=pk)
                serializer = StaffSalarySerializer(record, data=request.data, partial=True, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Salary record updated successfully!", "data": serializer.data}, status=200)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in StaffSalaryUpdateAPIView PUT: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class StaffSalaryMonthWiseAPIView(APIView):
    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching month-wise staff salaries for tenant: {tenant.name}")
                salaries = (
                    StaffSalary.objects
                    .annotate(month=TruncMonth('date'))
                    .values('staff__name', 'staff__contact', 'staff__address', 'month')
                    .annotate(total_salary=Sum('salary'))
                    .order_by('month')
                )
                return Response(salaries, status=200)
        except Exception as e:
            logger.error(f"Error in StaffSalaryMonthWiseAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class VehicleStatisticsAPIView(APIView):
    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching vehicle statistics for tenant: {tenant.name}")
                total_vehicles = Vehicle.objects.count()
                current_in_inventory = Vehicle.objects.filter(inventory_status='IN').count()
                total_outbound_vehicles = OutboundVehicle.objects.count()
                total_selling_price = OutboundVehicle.objects.aggregate(total=Sum('selling_price'))['total'] or 0
                total_purchase_price = Vehicle.objects.aggregate(total=Sum('purchase_price'))['total'] or 0
                total_maintenance_cost = MaintenanceRecord.objects.aggregate(total=Sum('cost'))['total'] or 0
                final_profit = total_selling_price - (total_purchase_price + total_maintenance_cost)

                data = {
                    "total_vehicles_added": total_vehicles,
                    "current_in_inventory": current_in_inventory,
                    "total_outbound_vehicles": total_outbound_vehicles,
                    "total_selling_price": total_selling_price,
                    "total_purchase_price": total_purchase_price,
                    "total_maintenance_cost": total_maintenance_cost,
                    "final_profit": final_profit,
                }

                return Response(data, status=200)
        except Exception as e:
            logger.error(f"Error in VehicleStatisticsAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

class InvoiceAPIView(APIView):
    def post(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Creating invoice for tenant: {tenant.name}")
                serializer = InvoiceSerializer(data=request.data, context={"request": request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({"message": "Invoice added successfully!", "data": serializer.data}, status=201)
                return Response(serializer.errors, status=400)
        except Exception as e:
            logger.error(f"Error in InvoiceAPIView POST: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)

    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"Session: {dict(request.session)}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}, User: {request.user}")
        logger.debug(f"Cookies: {request.COOKIES}")
        logger.debug(f"Session ID from cookies: {request.COOKIES.get('sessionid')}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                logger.debug(f"Fetching invoices for tenant: {tenant.name}")
                invoices = Invoice.objects.all()
                serializer = InvoiceSerializer(invoices, many=True, context={"request": request})
                return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in InvoiceAPIView GET: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)


class GetOutboundVehiclesAPIView(APIView):
    def get(self, request):
        logger.debug(f"Request headers: {request.headers}")
        logger.debug(f"User authenticated: {request.user.is_authenticated}")

        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt")
            return Response({"detail": "Authentication credentials were not provided."}, 
                           status=status.HTTP_401_UNAUTHORIZED)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}")

        try:
            with tenant_context(tenant):
                # Use select_related to include vehicle data in the same query
                outbound_vehicles = OutboundVehicle.objects.all().select_related('vehicle')
                serializer = OutboundVehicleSerializer(outbound_vehicles, many=True)
                return Response({
                    "message": "Outbound vehicles fetched successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error in GetOutboundVehiclesAPIView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            
class SalesStatsAPIView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            logger.warning("Unauthorized access attempt: No active session")
            return Response({"detail": "Authentication credentials were not provided."}, status=401)

        tenant = request.tenant
        logger.debug(f"Tenant: {tenant.name}, ID: {tenant.id}")

        try:
            with tenant_context(tenant):
                # Inventory stats
                total_vehicles = Vehicle.objects.count()
                total_in_inventory = Vehicle.objects.filter(inventory_status='IN').count()
                
                # Outbound stats
                outbound_vehicles = OutboundVehicle.objects.all()
                total_outbound = outbound_vehicles.count()
                total_sales_revenue = outbound_vehicles.aggregate(Sum('selling_price'))['selling_price__sum'] or 0
                total_other_expenses = outbound_vehicles.aggregate(Sum('other_expense'))['other_expense__sum'] or 0
                total_purchase_price = Vehicle.objects.aggregate(Sum('purchase_price'))['purchase_price__sum'] or 0
                
                # Monthly sales trend (last 12 months)
                now = timezone.now()
                monthly_sales = []
                for i in range(12):
                    start_date = now - timedelta(days=30 * (12 - i))
                    end_date = now - timedelta(days=30 * (11 - i))
                    monthly_total = outbound_vehicles.filter(
                        outbound_date__range=[start_date, end_date]
                    ).aggregate(Sum('selling_price'))['selling_price__sum'] or 0
                    monthly_sales.append(monthly_total)

                return Response({
                    "total_vehicles": total_vehicles,
                    "total_in_inventory": total_in_inventory,
                    "total_outbound": total_outbound,
                    "total_sales_revenue": total_sales_revenue,
                    "total_other_expenses": total_other_expenses,
                    "total_purchase_price": total_purchase_price,
                    "monthly_sales": monthly_sales
                }, status=200)
        except Exception as e:
            logger.error(f"Error in SalesStatsAPIView: {str(e)}", exc_info=True)
            return Response({"error": f"Server error: {str(e)}"}, status=500)